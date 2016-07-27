import csv
import os.path
import openpyxl


from os import listdir
from os.path import isfile, join
from sys import argv


def main():
	#generate appropriate path depending on location of code
	path = os.path.dirname(os.path.abspath(""))
	path = os.path.dirname(path)
	inpath = os.path.join(path,"data","raw")
	outputFile = os.path.join(path,"data","intermediate","weeklyOptionData.txt")
	years = (2006,2007,2008,2009,2010,2011,2012,2013,2014,2015,2016)

	#open output csv
	with open(outputFile,'w') as outPut:
		count = 1
		#loop through years
		for y in years:
			#loop through all files within year folders
			yearpath = os.path.join(inpath,str(y),str(y))
			for fn in os.listdir(yearpath):
				#create data dict pull data from file names and open file
				dataD = {}
				dataD[0] = (fn[6:8],"sheetType")
				wb = openpyxl.load_workbook(os.path.join(yearpath,fn))
				sheet = wb.active

				#pull all data from cells in dictionary via loop
				for i in range(2,21):
					dataD[i] = (sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=0).value)
				for j in range(22,37):
					dataD[j] = (sheet.cell(row=j,column=2).value,sheet.cell(row=j,column=0).value)
				#write header row if first iteration
				if count == 1:
					headerString = ""
					for value in dataD.values():
						headerString = headerString + str(value[1]) + "\t"
					outPut.write(headerString + "\n")
				# write line
				lineString = ""
				for value in dataD.values():
					lineString = lineString + str(value[0]) + "\t"
				outPut.write(lineString + "\n")
				count = count + 1
		# break test
			#if count > 5:
			#	break






if __name__ == "__main__":
	main()